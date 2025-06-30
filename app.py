import matplotlib
matplotlib.use('Agg')
from flask import Flask, request, send_file, jsonify, Response
import requests
import fitz  # PyMuPDF
import io
from PIL import Image
import pandas as pd
import matplotlib.pyplot as plt
import logging
import traceback
import tempfile
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 日志配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

@app.route('/PDFViewer')
def pdf_to_png():
    pdf_url = request.args.get('Url')
    page_index = request.args.get('PageIndex', default='1')
    logger.info(f"/PDFViewer called, Url={pdf_url}, PageIndex={page_index}")
    try:
        page_index = int(page_index)
    except Exception as e:
        logger.error(f"PageIndex参数转换失败: {e}")
        return Response('无效的参数', mimetype='text/plain')
    if not pdf_url:
        logger.error("缺少Url参数")
        return jsonify({'error': '缺少Url参数'}), 400
    try:
        resp = requests.get(pdf_url, timeout=15)
        logger.info(f"PDF下载状态: {resp.status_code}")
        if resp.status_code != 200:
            logger.error(f"PDF下载失败: {resp.status_code}")
            return jsonify({'error': 'PDF下载失败', 'status_code': resp.status_code}), 400
        pdf_bytes = resp.content
    except Exception as e:
        logger.error(f"PDF下载异常: {e}\n{traceback.format_exc()}")
        return jsonify({'error': f'PDF下载异常: {str(e)}'}), 400
    try:
        pdf_doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        page_count = pdf_doc.page_count
        logger.info(f"PDF总页数: {page_count}")
        if page_count == 0 or page_index < 1 or page_index > page_count:
            logger.error(f"无效的PageIndex: {page_index}")
            return Response('无效的参数', mimetype='text/plain')
        zoom = 300 / 72  # 300DPI
        mat = fitz.Matrix(zoom, zoom)
        page = pdf_doc[page_index - 1]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.open(io.BytesIO(pix.tobytes('png')))
        max_width = 1920
        if img.width > max_width:
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), Image.LANCZOS)
        img_io = io.BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        logger.info(f"PDF第{page_index}页渲染成功")
        return send_file(img_io, mimetype='image/png', as_attachment=False, download_name=f'page{page_index}.png')
    except Exception as e:
        logger.error(f"PDF处理异常: {e}\n{traceback.format_exc()}")
        return Response('无效的参数', mimetype='text/plain')

@app.route('/GetPDFPageCount')
def get_pdf_page_count():
    pdf_url = request.args.get('Url')
    logger.info(f"/GetPDFPageCount called, Url={pdf_url}")
    if not pdf_url:
        logger.error("缺少Url参数")
        return Response('0', mimetype='text/plain')
    try:
        resp = requests.get(pdf_url, timeout=15)
        logger.info(f"PDF下载状态: {resp.status_code}")
        if resp.status_code != 200:
            logger.error(f"PDF下载失败: {resp.status_code}")
            return Response('0', mimetype='text/plain')
        pdf_bytes = resp.content
        pdf_doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        count = pdf_doc.page_count
        logger.info(f"PDF总页数: {count}")
        return Response(str(count), mimetype='text/plain')
    except Exception as e:
        logger.error(f"PDF页数获取异常: {e}\n{traceback.format_exc()}")
        return Response('0', mimetype='text/plain')

@app.route('/ExcelViewer')
def excel_viewer():
    excel_url = request.args.get('Url')
    sheet_index = request.args.get('SheetIndex', default='1')
    logger.info(f"/ExcelViewer called, Url={excel_url}, SheetIndex={sheet_index}")
    try:
        sheet_index = int(sheet_index)
    except Exception as e:
        logger.error(f"SheetIndex参数转换失败: {e}")
        return Response('无效的参数', mimetype='text/plain')
    if not excel_url:
        logger.error("缺少Url参数")
        return Response('无效的参数', mimetype='text/plain')
    try:
        resp = requests.get(excel_url, timeout=15)
        logger.info(f"Excel下载状态: {resp.status_code}")
        if resp.status_code != 200:
            logger.error(f"Excel下载失败: {resp.status_code}")
            return Response('无效的参数', mimetype='text/plain')
        excel_bytes = io.BytesIO(resp.content)
        wb = load_workbook(excel_bytes, data_only=True)
        sheet_names = wb.sheetnames
        logger.info(f"Excel Sheet数量: {len(sheet_names)}")
        if sheet_index < 1 or sheet_index > len(sheet_names):
            logger.error(f"无效的SheetIndex: {sheet_index}")
            return Response('无效的参数', mimetype='text/plain')
        ws = wb[sheet_names[sheet_index-1]]
        # 获取合并单元格信息
        merged_ranges = ws.merged_cells.ranges
        merge_map = {}  # {(row, col): (rowspan, colspan)}
        skip_cells = set()
        for mr in merged_ranges:
            min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1
            merge_map[(min_row, min_col)] = (rowspan, colspan)
            for r in range(min_row, max_row+1):
                for c in range(min_col, max_col+1):
                    if not (r == min_row and c == min_col):
                        skip_cells.add((r, c))
        # 统计所有行
        all_rows = list(ws.iter_rows(values_only=True))
        # 统计有效列（有内容或为合并单元格起始的列）
        max_col = max((len(row) for row in all_rows), default=0)
        valid_cols = set()
        for i, row in enumerate(all_rows, 1):
            for j in range(1, max_col+1):
                # 有内容
                if j <= len(row) and row[j-1] not in (None, ''):
                    valid_cols.add(j)
                # 是合并单元格起始
                if (i, j) in merge_map:
                    valid_cols.add(j)
        valid_cols = sorted(valid_cols)
        # 生成HTML表格，只渲染有效列
        html = '<table>'
        for i, row in enumerate(all_rows, 1):
            html += '<tr>'
            col_idx = 1
            logical_col = 0
            while col_idx <= max_col:
                if col_idx not in valid_cols:
                    col_idx += 1
                    continue
                if (i, col_idx) in skip_cells:
                    col_idx += 1
                    continue
                val = ''
                if col_idx <= len(row):
                    cell = row[col_idx-1]
                    val = '' if cell is None else str(cell)
                attrs = ''
                colspan = 1
                if (i, col_idx) in merge_map:
                    rowspan, colspan = merge_map[(i, col_idx)]
                    if rowspan > 1:
                        attrs += f' rowspan="{rowspan}"'
                    if colspan > 1:
                        attrs += f' colspan="{colspan}"'
                tag = 'th' if i == 1 else 'td'
                html += f'<{tag}{attrs}>{val}</{tag}>'
                col_idx += colspan
            html += '</tr>'
        html += '</table>'
        # 包装完整HTML，设置样式，table-layout:fixed
        html_full = f"""
        <html>
        <head>
        <meta charset='utf-8'>
        <style>
        body {{ background: #fff; margin: 0; padding: 0; }}
        table {{ border-collapse: collapse; font-size: 16px; font-family: 'SimHei', 'Microsoft YaHei', Arial, sans-serif; table-layout: fixed; }}
        th, td {{ border: 1px solid #888; padding: 6px 12px; max-width: 400px; word-break: break-all; text-align: center; }}
        th {{ background: #f2f2f2; }}
        </style>
        </head>
        <body>{html}</body>
        </html>
        """
        # 用playwright渲染HTML并截图
        try:
            logger.info("开始用playwright渲染HTML并截图...")
            with tempfile.NamedTemporaryFile('w', delete=False, suffix='.html', encoding='utf-8') as f:
                f.write(html_full)
                html_path = f.name
            img_io = io.BytesIO()
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(f'file://{html_path}')
                page.wait_for_selector('table')
                element = page.query_selector('table')
                img_bytes = element.screenshot(type='png')
                img_io.write(img_bytes)
                img_io.seek(0)
                browser.close()
            logger.info(f"Excel Sheet{sheet_index}渲染成功（playwright，合并单元格支持），图片已生成")
            return send_file(img_io, mimetype='image/png', as_attachment=False, download_name=f'sheet{sheet_index}.png')
        except Exception as e:
            logger.error(f"playwright渲染异常: {e}\n{traceback.format_exc()}")
            return Response('无效的参数', mimetype='text/plain')
    except Exception as e:
        logger.error(f"Excel处理异常: {e}\n{traceback.format_exc()}")
        return Response('无效的参数', mimetype='text/plain')

@app.route('/GetExcelSheetCount')
def get_excel_sheet_count():
    excel_url = request.args.get('Url')
    logger.info(f"/GetExcelSheetCount called, Url={excel_url}")
    if not excel_url:
        logger.error("缺少Url参数")
        return Response('0', mimetype='text/plain')
    try:
        resp = requests.get(excel_url, timeout=15)
        logger.info(f"Excel下载状态: {resp.status_code}")
        if resp.status_code != 200:
            logger.error(f"Excel下载失败: {resp.status_code}")
            return Response('0', mimetype='text/plain')
        excel_bytes = io.BytesIO(resp.content)
        xls = pd.ExcelFile(excel_bytes)
        count = len(xls.sheet_names)
        logger.info(f"Excel Sheet数量: {count}")
        return Response(str(count), mimetype='text/plain')
    except Exception as e:
        logger.error(f"Excel Sheet数量获取异常: {e}\n{traceback.format_exc()}")
        return Response('0', mimetype='text/plain')

if __name__ == '__main__':
    logger.info('服务启动，监听0.0.0.0:5000')
    app.run(host='0.0.0.0', port=5000, debug=True) 